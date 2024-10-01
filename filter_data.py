import calendar
import csv
import json
import os
from copy import copy
from csv import DictWriter
from dataclasses import MISSING
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional, Literal

import openpyxl

import consts
from consts import output_folder, process, missing_folder, input_folder

output_folder.mkdir(exist_ok=True)
process.mkdir(exist_ok=True)
missing_folder.mkdir(exist_ok=True)

global_state = {}
global_state_file = process / "state_l.json"
if global_state_file.exists():
    global_state = json.load(global_state_file.open())

global_state_dict = {}
global_state_dict_file = process / "state_d.json"
if global_state_dict_file.exists():
    global_state_dict = json.load(global_state_dict_file.open())


def create_state_for_year(year: int) -> list[list[list[bool]]]:
    state = []
    for month in range(1, 13):
        month_state = []
        state.append(month_state)
        for day in range(1, calendar.monthrange(year, month)[1]):
            month_state.append([False for _ in range(24)])
    return state


def create_dict_state_for_year(year: int) -> dict[str, dict[dict[str, bool]]]:
    state = {}
    for month in range(1, 13):
        m_str = date(year, month, 1).strftime("%b").lower()
        month_state = state.setdefault(m_str, {})
        for day in range(1, calendar.monthrange(year, month)[1]):
            month_state.setdefault(day, {hour: False for hour in range(24)})
    return state


def get_datetime(datetime_: str) -> datetime:
    parse_format = ["%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M"]
    for format_ in parse_format:
        try:
            return datetime.strptime(datetime_, format_)
        except ValueError as e:
            date_, time_ = datetime_.split(" ")
            hour_, minute_ = time_.split(":")
            proper_hour = hour_.rjust(2, "0")
            fixed_dt = f"{date_} {proper_hour}:{minute_}"
            try:
                return datetime.strptime(fixed_dt, format_)
            except:
                continue


def write2csv(data: list[dict], fieldnames: list[str], orig_file_stem: str):
    output_path = output_folder.joinpath(orig_file_stem + "_auto_filtered.csv")
    writer = csv.DictWriter(output_path.open("w", encoding="utf-8"),
                            fieldnames)
    writer.writeheader()
    for line in data:
        writer.writerow(line)


def write2excel(data: list[dict], fieldnames: list[str], orig_file_stem: str):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write headers
    for col, header in enumerate(fieldnames, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Write data
    for row, item in enumerate(data, start=2):
        for col, key in enumerate(fieldnames, start=1):
            sheet.cell(row=row, column=col, value=item.get(key, ''))

    # Adjust column widths
    for col in range(1, len(fieldnames) + 1):
        max_length = 0
        column = openpyxl.utils.get_column_letter(col)
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    output_path = output_folder.joinpath(orig_file_stem + "_auto_filtered.xlsx")
    # Save the workbook
    workbook.save(filename=output_path)


def run_filter():
    # ITERATE THROUGH ALL CSV FILES AND COLLECT THE EARLIEST FOR EACH HOUR
    collected_years = set()
    for csv_file in input_folder.glob("*.csv"):
        print(csv_file)
        # COLLECT
        calendar_sorted_posts = {}
        # reader = csv.DictReader(csv_file.open(encoding="utf-8"))
        reader = csv.reader(csv_file.open(encoding="utf-8"))
        header = next(reader)
        line_number = 1
        while True:
            try:
                line = next(reader)
                lined = dict(zip(header, line))
                # print(lined)
                # print(line["发布时间"])
                dt = get_datetime(lined["发布时间"])
                c_year = calendar_sorted_posts.setdefault(dt.year, {})
                c_month = c_year.setdefault(dt.month, {})
                c_day = c_month.setdefault(dt.day, {})
                if dt.hour not in c_day:
                    c_day[dt.hour] = (dt, lined)
                elif dt < c_day[dt.hour][0]:
                    c_day[dt.hour] = (dt, lined)
                collected_years.add(dt.year)
                line_number += 1
            except StopIteration:
                break
            except Exception as e:
                print(line_number, e)
                continue
        # create state list if the year is missing
        for year in collected_years:
            if str(year) not in global_state:
                global_state[str(year)] = create_state_for_year(year)
                global_state_dict[str(year)] = (create_dict_state_for_year(year))

        fieldnames = ["date", "hour"] + list(header)
        fieldnames[2] = 'id'
        # print(fieldnames)
        rows2write: list[dict] = []
        for year, year_data in sorted(calendar_sorted_posts.items()):
            for month, month_data in sorted(year_data.items()):
                # print(f"month: {month}")
                for day, day_data in sorted(month_data.items()):
                    # print(f"day: {day}")
                    for hour, hour_data in sorted(day_data.items()):
                        global_state[str(year)][month - 1][day - 1][hour] = True
                        # print(f"hour: {hour}")
                        # print(hour_data)
                        dt, post = hour_data
                        post["hour"] = hour
                        post["date"] = post["发布时间"]
                        print(str(post['\ufeffid']))
                        pass
                        post['id'] = str(post['\ufeffid'])

                        del post['\ufeffid']
                        rows2write.append(post)

        if consts.outputformat == "csv":
            write2csv(rows2write, fieldnames, csv_file.stem)
        else:
            write2excel(rows2write, fieldnames, csv_file.stem)

    json.dump(global_state, global_state_file.open("w"))
    json.dump(global_state_dict, global_state_dict_file.open("w"))


def get_missing_for_year(year: int, month: Optional[int] = None) -> None:
    year_data = global_state[str(year)]
    with open(missing_folder / "all.txt", "w", encoding="utf-8") as fout:
        for month_idx, month_data in enumerate(year_data):
            if month and month_idx + 1 != month:
                continue
            for day_idx, day_data in enumerate(month_data):
                for hour_idx, collected in enumerate(day_data):
                    if not collected:
                        missing_hour = datetime(year, month_idx + 1, day_idx + 1, hour=hour_idx).strftime(
                            "%Y-%m-%d %H:%M")
                        print(missing_hour)
                        fout.write(missing_hour + os.linesep)


def get_missing_for_range(start: datetime, end: datetime) -> None:
    years = list(y + start.year for y in range(end.year - start.year + 1))
    print(years)
    cur_dt = copy(start)
    while cur_dt <= end:
        cur_dt += timedelta(hours=1)
        if not global_state[str(cur_dt.year)][cur_dt.month - 1][cur_dt.day - 1][cur_dt.hour]:
            print(cur_dt.strftime("%Y-%m-%d %H:%M"))


if __name__ == "__main__":
    run_filter()
    # get_missing_for_year(2023, 7)
    # get_missing_for_range(datetime(year=2023, month=7, day=1), datetime(year=2023, month=7, day=30))
